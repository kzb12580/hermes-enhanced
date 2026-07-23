"""
OfficeCLI 工具集 — 基于 OfficeCLI v1.0.140 的 Office 自动化
支持: Word (.docx) / Excel (.xlsx) / PowerPoint (.pptx)
特性: 原生OOXML动画、HTML/PNG渲染预览、实时预览服务器
"""
import os
import json
import subprocess
import sys
import logging
import tempfile
from pathlib import Path
from typing import Optional

_log = logging.getLogger(__name__)

# OfficeCLI 二进制路径（优先用户安装的，其次系统PATH）
_OFFICECLI_BIN = None

def _find_officecli() -> str:
    """查找 officecli 二进制路径"""
    global _OFFICECLI_BIN
    if _OFFICECLI_BIN:
        return _OFFICECLI_BIN
    
    # 1. 检查用户目录
    user_bin = Path.home() / ".officecli" / "officecli"
    if user_bin.exists():
        _OFFICECLI_BIN = str(user_bin)
        return _OFFICECLI_BIN
    
    # 2. 检查 npm 全局目录
    try:
        result = subprocess.run(
            ["npm", "root", "-g"],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode == 0:
            npm_global = Path(result.stdout.strip()) / "@officecli" / "officecli" / "bin" / "officecli"
            if npm_global.exists():
                _OFFICECLI_BIN = str(npm_global)
                return _OFFICECLI_BIN
    except Exception:
        pass
    
    # 3. 检查 PATH
    try:
        result = subprocess.run(
            ["which", "officecli"] if sys.platform != "win32" else ["where", "officecli"],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode == 0:
            _OFFICECLI_BIN = result.stdout.strip().split('\n')[0]
            return _OFFICECLI_BIN
    except Exception:
        pass
    
    # 4. 尝试直接调用（可能在PATH中）
    _OFFICECLI_BIN = "officecli"
    return _OFFICECLI_BIN


def _run_officecli(*args: str, timeout: int = 120) -> dict:
    """执行 officecli 命令并返回结果"""
    bin_path = _find_officecli()
    cmd = [bin_path] + list(args)
    
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=timeout,
            encoding="utf-8",
            errors="replace"
        )
        
        if result.returncode != 0:
            error_msg = result.stderr.strip() or result.stdout.strip() or "Unknown error"
            return {"error": error_msg, "success": False}
        
        # 尝试解析 JSON 输出
        output = result.stdout.strip()
        if output:
            try:
                return json.loads(output)
            except json.JSONDecodeError:
                return {"output": output, "success": True}
        
        return {"success": True}
        
    except subprocess.TimeoutExpired:
        return {"error": f"OfficeCLI timeout ({timeout}s)", "success": False}
    except FileNotFoundError:
        return {"error": "OfficeCLI not found. Install: npm install -g @officecli/officecli", "success": False}
    except Exception as e:
        return {"error": str(e), "success": False}


def _safe_path(p: str) -> str:
    """路径净化"""
    from tools.safe_file_ops import atomic_save, backup_file
    resolved = Path(p).expanduser().resolve()
    return str(resolved)


# ═══════════════════════════════════════════════════════════════════════════
# Word 工具
# ═══════════════════════════════════════════════════════════════════════════

def create_word(path: str, title: str = "", content: str = "", 
                template: str = "", font_size: int = 12, 
                line_spacing: float = 1.5) -> dict:
    """创建 Word 文档"""
    path = _safe_path(path)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    
    # 创建空文档（--force 覆盖）
    result = _run_officecli("create", path, "--force")
    if not result.get("success"):
        return result
    
    # 添加标题
    if title:
        _run_officecli("add", path, "/", "--type", "paragraph",
                       "--prop", f"text={title}", "--prop", "style=Heading1")
    
    # 添加内容
    if content:
        for para in content.split("\n"):
            if para.strip():
                _run_officecli("add", path, "/", "--type", "paragraph",
                              "--prop", f"text={para.strip()}")
    
    # 保存
    _run_officecli("save", path)
    
    return {"path": path, "success": True}


def read_word(path: str) -> dict:
    """读取 Word 文档内容"""
    path = _safe_path(path)
    if not os.path.isfile(path):
        return {"error": f"File not found: {path}", "success": False}
    
    # 使用 view text 命令读取
    result = _run_officecli("view", path, "text")
    if not result.get("success"):
        return result
    
    # 解析文本输出
    output = result.get("output", "")
    paragraphs = [line for line in output.split("\n") if line.strip()]
    
    return {"paragraphs": paragraphs, "success": True}


def edit_word(path: str, operations: list[dict]) -> dict:
    """编辑 Word 文档"""
    path = _safe_path(path)
    if not os.path.isfile(path):
        return {"error": f"File not found: {path}", "success": False}
    
    for i, op in enumerate(operations):
        t = op.get("type")
        if not t:
            return {"error": f"Op #{i} missing 'type'", "success": False}
        
        if t == "add_heading":
            level = op.get("level", 1)
            result = _run_officecli("add", f"{path}/paragraph[{op.get('index', -1)}]",
                                   "--type", f"heading{level}", "--text", op.get("text", ""))
        elif t == "add_paragraph":
            result = _run_officecli("add", f"{path}/paragraph[{op.get('index', -1)}]",
                                   "--type", "body", "--text", op.get("text", ""))
        elif t == "replace":
            # 使用 set 修改文本
            result = _run_officecli("set", f"{path}/paragraph[{op.get('index', 0)}]",
                                   "--text", op.get("new", ""))
        elif t == "add_image":
            img_path = op.get("image_path", "")
            if img_path:
                result = _run_officecli("add", f"{path}",
                                       "--type", "image", "--path", img_path)
        else:
            continue
        
        if not result.get("success"):
            return result
    
    return {"path": path, "operations": len(operations), "success": True}


# ═══════════════════════════════════════════════════════════════════════════
# Excel 工具
# ═══════════════════════════════════════════════════════════════════════════

def create_excel(path: str, sheets: list[dict]) -> dict:
    """创建 Excel 表格（智能引擎选择）"""
    path = _safe_path(path)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    
    # 计算总数据量
    total_rows = sum(len(s.get("data", [])) for s in sheets)
    total_cells = total_rows * max((len(s.get("headers", [])) for s in sheets), default=0)
    
    # 大数据量（>100行）使用 openpyxl 原生引擎
    if total_rows > 100:
        return _create_excel_native(path, sheets)
    
    # 小数据量使用 OfficeCLI
    return _create_excel_cli(path, sheets)


def _create_excel_native(path: str, sheets: list[dict]) -> dict:
    """使用 openpyxl 创建 Excel（高性能）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        
        for i, sheet_data in enumerate(sheets):
            ws = wb.active if i == 0 else wb.create_sheet()
            ws.title = sheet_data.get("name", f"Sheet{i+1}")
            
            headers = sheet_data.get("headers", [])
            data = sheet_data.get("data", [])
            
            # 写入表头
            if headers:
                for j, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=j, value=header)
                    cell.font = Font(color="FFFFFF", bold=True, size=11)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                start_row = 2
            else:
                start_row = 1
            
            # 写入数据
            for r, row_data in enumerate(data, start_row):
                for c, value in enumerate(row_data, 1):
                    ws.cell(row=r, column=c, value=value)
        
        path = _safe_path(path)
        wb.save(path)
        return {"path": path, "sheets": len(sheets), "engine": "openpyxl", "success": True}
    except Exception as e:
        return {"error": str(e), "success": False}


def _create_excel_cli(path: str, sheets: list[dict]) -> dict:
    """使用 OfficeCLI 创建 Excel（小数据量）"""
    path = _safe_path(path)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    
    result = _run_officecli("create", path, "--force")
    if not result.get("success"):
        return result
    
    for sheet_data in sheets:
        sheet_name = sheet_data.get("name", "Sheet1")
        headers = sheet_data.get("headers", [])
        data = sheet_data.get("data", [])
        
        _run_officecli("add", path, "/", "--type", "sheet", "--prop", f"name={sheet_name}")
        
        if headers:
            for col, header in enumerate(headers):
                col_letter = chr(65 + col)
                _run_officecli("set", path, f"/{sheet_name}/{col_letter}1",
                              "--prop", f"value={header}")
        
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                col_letter = chr(65 + col_idx)
                row_num = row_idx + 2
                _run_officecli("set", path, f"/{sheet_name}/{col_letter}{row_num}",
                              "--prop", f"value={value}")
    
    _run_officecli("save", path)
    return {"path": path, "sheets": len(sheets), "engine": "officecli", "success": True}


def read_excel(path: str, sheet_name: str = "", max_rows: int = 1000) -> dict:
    """读取 Excel 文件"""
    path = _safe_path(path)
    if not os.path.isfile(path):
        return {"error": f"File not found: {path}", "success": False}
    
    # 使用 view text 命令读取
    result = _run_officecli("view", path, "text")
    if not result.get("success"):
        return result
    
    # 解析表格输出
    output = result.get("output", "")
    sheets = {}
    current_sheet = None
    
    for line in output.split("\n"):
        if line.startswith("=== Sheet:"):
            current_sheet = line.split(":")[1].strip().rstrip(" =")
            if current_sheet.startswith("Sheet") and current_sheet[5:].isdigit():
                # 跳过默认空Sheet
                current_sheet = None
                continue
            sheets[current_sheet] = []
        elif current_sheet and line.strip():
            # 解析行数据: [/SheetName/row[N]] A1=val\tB1=val\t...
            if "\t" in line:
                row_data = {}
                for cell in line.split("\t"):
                    cell = cell.strip()
                    if "=" in cell:
                        ref, val = cell.split("=", 1)
                        row_data[ref] = val
                if row_data:
                    sheets[current_sheet].append(row_data)
    
    return {"sheets": sheets, "success": True}


def edit_excel(path: str, operations: list[dict]) -> dict:
    """编辑 Excel 文件"""
    path = _safe_path(path)
    if not os.path.isfile(path):
        return {"error": f"File not found: {path}", "success": False}
    
    for i, op in enumerate(operations):
        t = op.get("type")
        sheet = op.get("sheet", "Sheet1")
        
        if t == "set_cell":
            row = op.get("row", 1)
            col = op.get("col", 1)
            value = str(op.get("value", ""))
            result = _run_officecli("set", f"{path}/{sheet}/cell[{row},{col}]",
                                   "--value", value)
        elif t == "set_range":
            data = op.get("data", [])
            sr, sc = op.get("start_row", 1), op.get("start_col", 1)
            for r, row_data in enumerate(data):
                for c, val in enumerate(row_data):
                    _run_officecli("set", f"{path}/{sheet}/cell[{sr+r},{sc+c}]",
                                  "--value", str(val))
        elif t == "add_sheet":
            name = op.get("name", f"Sheet{i+1}")
            result = _run_officecli("add", f"{path}/sheet", "--name", name)
        elif t == "add_formula":
            row = op.get("row", 1)
            col = op.get("col", 1)
            formula = op.get("formula", "")
            result = _run_officecli("set", f"{path}/{sheet}/cell[{row},{col}]",
                                   "--formula", formula)
        else:
            continue
        
        if not result.get("success"):
            return result
    
    return {"path": path, "operations": len(operations), "success": True}


# ═══════════════════════════════════════════════════════════════════════════
# PowerPoint 工具 — 支持动画！
# ═══════════════════════════════════════════════════════════════════════════

def create_ppt(path: str, slides: Optional[list[dict]] = None, 
               layout: str = "16x9", title: str = "", author: str = "",
               slides_file: str = "") -> dict:
    """创建 PPT（batch 高性能模式）"""
    path = _safe_path(path)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    
    if slides_file:
        sf = _safe_path(slides_file)
        if not os.path.isfile(sf):
            return {"error": f"slides_file not found: {sf}", "success": False}
        with open(sf, "r", encoding="utf-8-sig") as f:
            slides = json.load(f)
    
    if not slides or not isinstance(slides, list):
        return {"error": "slides required", "success": False}
    
    result = _run_officecli("create", path, "--force")
    if not result.get("success"):
        return result
    
    # 构建 batch 命令
    batch = []
    if title:
        batch.append({"command": "set", "path": "/", "props": {"title": title}})
    if author:
        batch.append({"command": "set", "path": "/", "props": {"author": author}})
    
    for slide_idx, slide in enumerate(slides):
        slide_path = f"/slide[{slide_idx + 1}]"
        batch.append({"command": "add", "parent": "/", "type": "slide"})
        
        bg = slide.get("background", {})
        if bg.get("color"):
            batch.append({"command": "set", "path": slide_path, "props": {"bgColor": bg["color"]}})
        
        for elem_idx, elem in enumerate(slide.get("elements", [])):
            elem_type = elem.get("type", "text")
            if elem_type == "text":
                props = {"text": elem.get("text", ""),
                         "x": str(elem.get("x", 1)), "y": str(elem.get("y", 1)),
                         "w": str(elem.get("w", 8)), "h": str(elem.get("h", 2))}
                if elem.get("fontSize"): props["fontSize"] = str(elem["fontSize"])
                if elem.get("color"): props["color"] = elem["color"]
                if elem.get("bold"): props["bold"] = "true"
                batch.append({"command": "add", "parent": slide_path, "type": "textbox", "props": props})
            elif elem_type == "shape":
                props = {"x": str(elem.get("x", 0)), "y": str(elem.get("y", 0)),
                         "w": str(elem.get("w", 1)), "h": str(elem.get("h", 1))}
                if elem.get("fill", {}).get("color"): props["fillColor"] = elem["fill"]["color"]
                batch.append({"command": "add", "parent": slide_path, "type": elem.get("shape", "rect"), "props": props})
            elif elem_type == "image" and elem.get("path"):
                props = {"path": elem["path"], "x": str(elem.get("x", 0)), "y": str(elem.get("y", 0)),
                         "w": str(elem.get("w", 4)), "h": str(elem.get("h", 3))}
                batch.append({"command": "add", "parent": slide_path, "type": "picture", "props": props})
        
        for anim in slide.get("animations", []):
            props = {"effect": anim.get("effect", "fade"), "class": anim.get("class", "entrance"),
                     "duration": str(anim.get("duration", 500)), "trigger": anim.get("trigger", "onClick")}
            if anim.get("direction"): props["direction"] = anim["direction"]
            if anim.get("delay"): props["delay"] = str(anim["delay"])
            batch.append({"command": "add", "parent": f"{slide_path}/shape[{anim.get('shape_index', 0) + 1}]",
                         "type": "animation", "props": props})
    
    # 执行 batch
    if batch:
        import subprocess
        import tempfile
        try:
            # 写入临时文件
            with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, 
                                           dir=os.path.dirname(path) or '.') as f:
                json.dump(batch, f, ensure_ascii=False)
                batch_file = f.name
            
            proc = subprocess.run(["officecli", "batch", path, "--input", batch_file],
                                  capture_output=True, text=True, timeout=120)
            
            os.unlink(batch_file)
            
            if proc.returncode != 0:
                return {"error": f"batch failed: {proc.stderr}", "success": False}
        except Exception as e:
            return {"error": str(e), "success": False}
    
    _run_officecli("save", path)
    return {"path": path, "slides": len(slides), "success": True}


def add_ppt_animation(path: str, slide_index: int, shape_index: int,
                      effect: str, anim_class: str = "entrance",
                      duration: int = 500, trigger: str = "onClick",
                      direction: str = "", delay: int = 0) -> dict:
    """
    为 PPT 元素添加动画
    
    参数:
      - path: PPT文件路径
      - slide_index: 幻灯片索引（从1开始）
      - shape_index: 形状索引（从1开始）
      - effect: 动画效果 (appear/fade/fly/zoom/wipe/bounce/spin/grow等)
      - anim_class: 动画类别 (entrance/exit/emphasis/motion)
      - duration: 持续时间（毫秒）
      - trigger: 触发方式 (onClick/withPrevious/afterPrevious)
      - direction: 方向 (in/out/left/right/up/down)
      - delay: 延迟（毫秒）
    """
    path = _safe_path(path)
    if not os.path.isfile(path):
        return {"error": f"File not found: {path}", "success": False}
    
    # 先打开文件
    _run_officecli("open", path)
    
    try:
        anim_path = f"/slide[{slide_index}]/shape[{shape_index}]"
        
        cmd_args = ["add", path, anim_path, "--type", "animation",
                    "--prop", f"effect={effect}",
                    "--prop", f"class={anim_class}",
                    "--prop", f"duration={duration}",
                    "--prop", f"trigger={trigger}"]
        
        if direction:
            cmd_args.extend(["--prop", f"direction={direction}"])
        if delay:
            cmd_args.extend(["--prop", f"delay={delay}"])
        
        result = _run_officecli(*cmd_args)
        
        # 保存文件
        _run_officecli("save", path)
        
        return result
    except Exception as e:
        _run_officecli("save", path)
        return {"error": str(e), "success": False}


# ═══════════════════════════════════════════════════════════════════════════
# 渲染预览工具 — OfficeCLI 的另一个强项
# ═══════════════════════════════════════════════════════════════════════════

def render_office(path: str, output: str = "", format: str = "html",
                  slide: int = 0) -> dict:
    """
    渲染 Office 文档为 HTML 或 PNG 预览
    
    参数:
      - path: Office 文件路径
      - output: 输出路径（可选，默认与输入同名）
      - format: 输出格式 (html/png/screenshot)
      - slide: PPT指定幻灯片（0=全部）
    """
    path = _safe_path(path)
    if not os.path.isfile(path):
        return {"error": f"File not found: {path}", "success": False}
    
    if not output:
        base = os.path.splitext(path)[0]
        ext = "html" if format == "html" else "png"
        output = f"{base}.{ext}"
    
    output = _safe_path(output)
    os.makedirs(os.path.dirname(output) or ".", exist_ok=True)
    
    # 使用 view 命令渲染
    if format == "html":
        cmd_args = ["view", path, "html", "-o", output]
    elif format in ("png", "screenshot"):
        cmd_args = ["view", path, "screenshot", "-o", output]
        if slide:
            cmd_args.extend(["--page", str(slide)])
    else:
        return {"error": f"Unsupported format: {format}", "success": False}
    
    result = _run_officecli(*cmd_args)
    if result.get("success"):
        result["output"] = output
    
    return result


def get_office_info(path: str) -> dict:
    """获取 Office 文档结构信息"""
    path = _safe_path(path)
    if not os.path.isfile(path):
        return {"error": f"File not found: {path}", "success": False}
    
    return _run_officecli("dump", path, "--format", "outline")


def validate_ppt(path: str) -> dict:
    """验证 PPT 结构完整性"""
    path = _safe_path(path)
    if not os.path.isfile(path):
        return {"error": f"File not found: {path}", "success": False}
    
    return _run_officecli("validate", path)


# ═══════════════════════════════════════════════════════════════════════════
# 实时预览服务器
# ═══════════════════════════════════════════════════════════════════════════

_preview_process = None

def start_preview_server(port: int = 26315) -> dict:
    """启动 OfficeCLI 实时预览服务器"""
    global _preview_process
    
    if _preview_process and _preview_process.poll() is None:
        return {"status": "running", "port": port, "success": True}
    
    try:
        import subprocess
        _preview_process = subprocess.Popen(
            ["officecli", "preview", "--port", str(port)],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE
        )
        return {"status": "started", "port": port, "url": f"http://localhost:{port}", "success": True}
    except Exception as e:
        return {"error": str(e), "success": False}


def stop_preview_server() -> dict:
    """停止预览服务器"""
    global _preview_process
    
    if _preview_process:
        _preview_process.terminate()
        _preview_process = None
        return {"status": "stopped", "success": True}
    
    return {"status": "not_running", "success": True}
